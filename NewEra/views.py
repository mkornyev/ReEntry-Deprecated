
###################################### 
#              IMPORTS 
######################################

import ast
import os
# Used to export to Excel spreadsheet
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from datetime import datetime

from django.http import Http404, HttpResponse, HttpResponseRedirect #, JsonResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from django.views.decorators.csrf import ensure_csrf_cookie
from django.db.models import Q
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger

from django.contrib.auth.decorators import login_required
from django.contrib.auth import authenticate
from django.contrib.auth import login as auth_login # 'login' can't clash w/view names in namespace 
from django.contrib.auth import logout as auth_logout # 'logout' can't clash w/view names in namespace 
from django.contrib import messages

from django.utils import timezone

from NewEra.models import User, CaseLoadUser, Resource, Referral, Tag
from NewEra.forms import LoginForm, RegistrationForm, EditUserForm, EditSelfUserForm, CaseLoadUserForm, CreateResourceForm, TagForm, ResourceFilter, EditReferralNotesForm


###################################### 
#              CONSTANTS 
######################################

# Define the number of referrals or paginations shown on a page
RESOURCE_PAGINATION_COUNT = 30
REFERRAL_PAGINATION_COUNT = 20

###################################### 
#              MISC ACTIONS  
######################################

def home(request): 
	markReferralAsSeen(request)
	return render(request, 'NewEra/home.html', {})

def login(request):
	if request.user.is_authenticated:
		return redirect(reverse('Home'))

	context = {}

	if request.method == 'GET':
		context['form'] = LoginForm()
		return render(request, 'NewEra/login.html', context)

	form = LoginForm(request.POST)
	context['form'] = form

	if not form.is_valid():
		return render(request, 'NewEra/login.html', context)

	user = authenticate(username=form.cleaned_data['username'],
							password=form.cleaned_data['password'])

	auth_login(request, user)
	messages.success(request, 'Logged in as {} {}.'.format(user.first_name, user.last_name))
	return redirect(reverse('Home'))

@login_required
def logout(request):
	auth_logout(request)
	messages.success(request, 'Successfully logged out.')
	return redirect(reverse('Login'))

def about(request):
	return render(request, 'NewEra/about.html')

# Function to update the referral given a GET request (to a resource) with a querystring timestamp
def markReferralAsSeen(request):
	if 'key' not in request.GET:
		return 

	try:
		keyDate = datetime.strptime(request.GET['key'], '%Y-%m-%d %H:%M:%S.%f')
	except: 
		return 

	referrals = Referral.objects.filter(referral_date=keyDate)
	
	if referrals.count() == 1:
		referral = referrals.first()
		referral.date_accessed = datetime.now()
		referral.save()

# Function to check visitor cookie, and see if they accessed the resource
def isUniqueVisit(request, response, id): 
	siteStaff = request.COOKIES.get('siteStaff', '')

	if request.user.is_authenticated or siteStaff == 'true':
		response.set_cookie('siteStaff', 'true')
		return False 

	visitedResources = request.COOKIES.get('visitedResources', '').split(';')

	if visitedResources == ['']:
		response.set_cookie('visitedResources', str(id))
		return True 
	elif str(id) in visitedResources:
		return False
	else: 
		val = ';'.join(visitedResources)
		val = val + ';' + str(id)
		response.set_cookie('visitedResources', val)
		return True 
	
	return False 

###################################### 
#              RESOURCES  
######################################

def resources(request):
	all_resources = Resource.objects.all()
	context = { 'filter': ResourceFilter(request.GET, queryset=all_resources) }

	if request.method == 'GET':
		# SEARCH QUERY
		query = request.GET.get('query')

		# Set context value based on if the filter is set on the resources page
		if query:
			context['active_resources'] = context['filter'].qs.filter( Q(is_active=True) & (Q(name__icontains=query) | Q(description__icontains=query)) )
			context['inactive_resources'] = context['filter'].qs.filter( Q(is_active=False) & (Q(name__icontains=query) | Q(description__icontains=query)) )
		else: 
			context['active_resources'] = context['filter'].qs.filter(is_active=True)
			context['inactive_resources'] = context['filter'].qs.filter(is_active=False)

		# FILTER QUERY - build a query param for use in pagination links
		filterParams = request.GET.getlist('tags', '')
		filterQueryString = ''

		if filterParams: 
			for id in filterParams:
				filterQueryString += '&tags='
				filterQueryString += id

		context['filterQuery'] = filterQueryString

		# PAGINATION
		active_page = request.GET.get('a_page', 1)
		inactive_page = request.GET.get('i_page', 1)
		paginator = Paginator(context['active_resources'], RESOURCE_PAGINATION_COUNT)
		inactive_paginator = Paginator(context['inactive_resources'], RESOURCE_PAGINATION_COUNT)
		
		# Render the user's selected page for active resources
		try:
			activeResources = paginator.page(active_page)
		except PageNotAnInteger:
			activeResources = paginator.page(1)
		except EmptyPage:
			activeResources = paginator.page(paginator.num_pages)

		# Render the user's selected page for inactive resources
		try:
			inactiveResources = inactive_paginator.page(inactive_page)
		except PageNotAnInteger:
			inactiveResources = inactive_paginator.page(1)
		except EmptyPage:
			inactiveResources = inactive_paginator.page(inactive_paginator.num_pages)

		context['active_resources'] = activeResources
		context['inactive_resources'] = inactiveResources

	return render(request, 'NewEra/resources.html', context)

def get_resource(request, id):
	resource = get_object_or_404(Resource, id=id)
	context = { 'resource': resource, 'tags': resource.tags.all() }
	response = render(request, 'NewEra/get_resource.html', context)
	
	# Update the resource clicks
	if isUniqueVisit(request, response, id):
		resource.clicks = resource.clicks + 1
		resource.save()

	markReferralAsSeen(request)

	return response

# ***** Note about images *****
# They are uploaded to the system as type .JPEG or .PNG etc.
# And then saved as type django.FileField() 
# *****************************
def get_resource_image(request, id): 
	resource = get_object_or_404(Resource, id=id)

	if not resource.image:
		raise Http404

	return HttpResponse(resource.image, content_type=resource.content_type)

# Deletes the given image if it exists
@login_required
def deleteImage(request, oldImage):
	if oldImage: 
		BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
		IMAGE_ROOT = os.path.join(BASE_DIR, 'NewEra/user_uploads/' + oldImage.name)
		os.remove(IMAGE_ROOT)

@login_required
def create_resource(request):
	context = {}
	form = CreateResourceForm()
	context['form'] = form
	context['action'] = 'Create'

	if request.method == 'POST':
		resource = Resource()
		form = CreateResourceForm(request.POST, request.FILES, instance=resource)
		
		if form.is_valid():
			# Update the Resource content_type MANUALLY 
			pic = form.cleaned_data['image']
			if pic and pic != '':
				print('Uploaded image: {} (type={})'.format(pic, type(pic)))
				resource.content_type = form.cleaned_data['image'].content_type

			form.save()
			resource.save()

			messages.success(request, 'Resource successfully created!')

			return redirect('Resources')
	else:
		form = CreateResourceForm()

	return render(request, 'NewEra/edit_resource.html', context)

@login_required
def edit_resource(request, id):
	resource = get_object_or_404(Resource, id=id)
	oldImage = resource.image

	if request.method == "POST":
		form = CreateResourceForm(request.POST, request.FILES, instance=resource)
    
		if form.is_valid():

			pic = form.cleaned_data['image']
			if pic and pic != '':
				
				# Update content type & remove old image
				try: 
					# THE FOLLOWING LINE MAY THROW:
					# After being set, the resource image is saved as a FileField type (and not an Image)
					# As a result, it will not have a content_type attribute the way image files will

					resource.content_type = form.cleaned_data['image'].content_type
					deleteImage(request, oldImage)
				except: 
					pass

			form.save()
			resource.save()

			messages.success(request, '{} successfully edited.'.format(resource.name))
			return redirect('Show Resource', id=resource.id)
	else:
		form = CreateResourceForm(instance=resource)
	return render(request, 'NewEra/edit_resource.html', {'form': form, 'resource': resource, 'action': 'Edit'})

@login_required
def delete_resource(request, id):
	resource = get_object_or_404(Resource, id=id)

	if request.method == 'POST':
		# Delete the resource (and its image) only if it has never been referred
		if (resource.referrals.count() == 0):
			deleteImage(request, resource.image)
			resource.delete()
			messages.success(request, '{} successfully deleted.'.format(resource.name))
			return redirect('Resources')
		# Otherwise, deactivate the resource
		else:
			resource.is_active = False
			resource.save()
			messages.success(request, '{} was made inactive.'.format(resource.name))
			return redirect('Show Resource', id=resource.id)
	return render(request, 'NewEra/delete_resource.html', {'resource': resource})

@login_required
def resetViews(request):
	if request.method == 'POST':
		Resource.objects.all().update(clicks=0)
		messages.success(request, 'Reset all resource views')
		return redirect(reverse('Dashboard'))
	return render(request, 'NewEra/reset_view_counts.html', {})

###################################### 
#              REFERRALS  
######################################

@login_required
def create_referral(request):
	resources = request.GET.get('resources', None)	
	recipients = []
	
	if request.user.is_superuser: 
		recipients = CaseLoadUser.objects.filter(is_active=True).all()
	elif request.user.is_staff: 
		recipients = CaseLoadUser.objects.filter(is_active=True).filter(user=request.user)

	if request.method == 'GET' and resources:
		resources = [digit.strip() for digit in ast.literal_eval(resources)] # Safely parse array
		resources = [ get_object_or_404(Resource, id=resourceId) for resourceId in resources ]

		return render(request, 'NewEra/create_referral.html', {'resources': resources, 'recipients': recipients })

	elif request.method == 'POST': 
		phoneInput = ''.join(digit for digit in request.POST.get('phone', '') if digit.isdigit())
		
		# Referral to someone on a case load
		if 'resources[]' in request.POST and 'user_id' in request.POST and 'notes' in request.POST: 
			caseload_user = get_object_or_404(CaseLoadUser, id=request.POST['user_id'])
			resources = [get_object_or_404(Resource, id=num) for num in request.POST.getlist('resources[]')]

			# Change the message introduction depending on whether the case load user has a nickname or not
			if caseload_user.nickname:
				nameInput = caseload_user.nickname
			else:
				nameInput = caseload_user.first_name
			referral = Referral(email=caseload_user.email, phone=caseload_user.phone, notes=request.POST['notes'], user=request.user, caseUser=caseload_user)

		# Out of system referral
		elif 'resources[]' in request.POST and 'phone' in request.POST and 'email' in request.POST and 'notes' in request.POST and (len(phoneInput) == 10 or len(phoneInput) == 0): 
			resources = [get_object_or_404(Resource, id=num) for num in request.POST.getlist('resources[]')]
			referral = Referral(email=request.POST['email'], phone=phoneInput, notes=request.POST['notes'], user=request.user)
			nameInput = request.POST['name']
			
		else: 
			messages.error(request, 'Please fill out all fields.')
			return render(request, 'NewEra/create_referral.html', {'resources': resources, 'recipients': recipients })
		
		referral.save()

		for r in resources: 
			referral.resource_set.add(r)
		
		# Send the referral via email and SMS
		referralTimeStamp = str(referral.referral_date)
		referral.sendEmail(referralTimeStamp, nameInput)
		referral.sendSMS(referralTimeStamp, nameInput)

	messages.success(request, 'Successfully created a new referral.')

	return redirect(reverse('Resources'))

@login_required
def referrals(request):
	# Show admins all referrals, and SOWs only their referrals
	if (request.user.is_superuser):
		referrals = Referral.objects.all().order_by('-referral_date')
	elif (request.user.is_staff):
		referrals = Referral.objects.all().filter(user=request.user).order_by('-referral_date')

	# PAGINATION
	page = request.GET.get('page', 1)
	paginator = Paginator(referrals, REFERRAL_PAGINATION_COUNT)
	
	# Render the user's selected page for referrals
	try:
		referrals = paginator.page(page)
	except PageNotAnInteger:
		referrals = paginator.page(1)
	except EmptyPage:
		referrals = paginator.page(paginator.num_pages)

	context = { 'referrals': referrals }

	return render(request, 'NewEra/referrals.html', context)

@login_required
def get_referral(request, id):
	referral = get_object_or_404(Referral, id=id)
	context = { 'referral': referral, 'resources': Resource.objects.all().filter(referrals=referral) }
	return render(request, 'NewEra/get_referral.html', context)

def edit_referral_notes(request, id):
	referral = get_object_or_404(Referral, id=id)

	if request.method == "POST":
		form = EditReferralNotesForm(request.POST, instance=referral)
    
		if form.is_valid():

			form.save()
			referral.save()

			messages.success(request, 'Referral notes updated.')
			return redirect('Show Referral', id=referral.id)
	else:
		form = EditReferralNotesForm(instance=referral)
	return render(request, 'NewEra/edit_referral_notes.html', {'form': form, 'referral': referral, 'action': 'Edit'})

###################################### 
#              CASE LOAD  
######################################

@login_required
def case_load(request):
	users = [] 
	context = {} 

	# Set users to show differently based on the current user logged in
	if request.user.is_superuser: 
		users = CaseLoadUser.objects.all()	
		# Changed to account for inactive users
		context['staff'] = User.objects.filter(is_active=True).order_by('first_name', 'last_name')
	elif request.user.is_staff:
		users = CaseLoadUser.objects.filter(user=request.user).order_by('first_name', 'last_name')
	else:  
		raise Http404

	if request.method == 'POST' and 'staff_id' in request.POST:
		staff_user = get_object_or_404(User, id=request.POST['staff_id'])
		load_user = CaseLoadUser(user=staff_user)
		form = CaseLoadUserForm(request.POST, instance=load_user)

		if not form.is_valid():
			context['form'] = form 
			context['caseload_users'] = users
			context['modalStatus'] = 'show'
			return render(request, 'NewEra/case_load.html', context)

		form.save()
		load_user.save() 
		messages.success(request, 'Successfully added {} {} to the CaseLoad.'.format(load_user.first_name, load_user.last_name))

	context['caseload_users'] = users
	context['form'] = CaseLoadUserForm()
	return render(request, 'NewEra/case_load.html', context)

@login_required
def get_case_load_user(request, id):
	case_load_user = get_object_or_404(CaseLoadUser, id=id)
	context = { 'case_load_user': case_load_user }
	return render(request, 'NewEra/get_case_load_user.html', context)

@login_required
def edit_case_load_user(request, id):
	case_load_user = get_object_or_404(CaseLoadUser, id=id)

	if request.method == "POST":
		form = CaseLoadUserForm(request.POST, instance=case_load_user)
    
		if form.is_valid():

			form.save()
			case_load_user.save()

			messages.success(request, '{} successfully edited.'.format(case_load_user.get_full_name()))
			return redirect('Show Case Load User', id=case_load_user.id)
	else:
		form = CaseLoadUserForm(instance=case_load_user)
	return render(request, 'NewEra/edit_case_load_user.html', {'form': form, 'case_load_user': case_load_user, 'action': 'Edit'})

@login_required
def delete_case_load_user(request, id):
	case_load_user = get_object_or_404(CaseLoadUser, id=id)

	if request.method == 'POST':
		# Delete the case load user only if they have never been referred
		if (case_load_user.get_referrals().count() == 0):
			case_load_user.delete()
			messages.success(request, '{} successfully deleted.'.format(case_load_user.get_full_name()))
			return redirect('Case Load')
		# Otherwise, delete the case load user
		else:
			case_load_user.is_active = False
			case_load_user.save()
			messages.success(request, '{} was made inactive.'.format(case_load_user.get_full_name()))
			return redirect('Show Case Load User', id=case_load_user.id)
	return render(request, 'NewEra/delete_case_load_user.html', {'case_load_user': case_load_user})

###################################### 
#              USER MANAGEMENT  
######################################

@login_required
def dashboard(request): 
	if not request.user.is_superuser:
		raise Http404

	admins = User.objects.filter(is_superuser=True)
	sows = User.objects.filter(is_superuser=False).filter(is_staff=True)
	context = {'admins':admins, 'sows':sows, 'form': RegistrationForm()}

	if request.method == 'POST':
		form = RegistrationForm(request.POST)
		context['form'] = form

		if not form.is_valid():
			context['modalStatus'] = 'show'
			return render(request, 'NewEra/dashboard.html', context)

		user = User.objects.create_user(username=form.cleaned_data['username'], 
										password=form.cleaned_data['password'],
										email=form.cleaned_data['email'],
										phone=form.cleaned_data['phone'],
										first_name=form.cleaned_data['first_name'],
										last_name=form.cleaned_data['last_name'])
		user.is_staff = True 
		user.is_superuser = False

		# Radio button input 
		if 'user_type' in request.POST and request.POST['user_type'] == 'admin': 
			user.is_superuser = True

		user.save()
		messages.success(request, 'Added a new user to the system.')
	
	context['form'] = RegistrationForm()
	return render(request, 'NewEra/dashboard.html', context)

@login_required
def edit_user(request, id):
	user = get_object_or_404(User, id=id)

	if request.method == "POST":
		# If a user is editing themselves, get the form for only that user
		if user == request.user:
			form = EditSelfUserForm(request.POST, instance=user)
		# If an admin is editing someone else, get the appropriate form
		else:
			form = EditUserForm(request.POST, instance=user)
    
		if form.is_valid():

			form.save()
			user.save()

			messages.success(request, '{} successfully edited.'.format(str(user)))
			return redirect('Dashboard')
	else:
		# If a user is editing themselves, get the form for only that user
		if user == request.user:
			form = EditSelfUserForm(instance=user)
		# If an admin is editing someone else, get the appropriate form
		else:
			form = EditUserForm(instance=user)
	return render(request, 'NewEra/edit_user.html', {'form': form, 'user': user, 'action': 'Edit'})

@login_required
def delete_user(request, id):
	user = get_object_or_404(User, id=id)

	if request.method == 'POST':
		# Delete the user only if they have never made a referral and there is no one on their case load
		if (user.get_referrals().count() == 0 and user.get_case_load().count() == 0):
			user.delete()
			messages.success(request, 'User successfully deleted.')
			return redirect('Dashboard')
		# Otherwise, deactivate the user
		else:
			user.is_active = False
			user.save()
			messages.success(request, '{} was made inactive.'.format(user.get_full_name()))
			return redirect('Dashboard')
	return render(request, 'NewEra/delete_user.html', {'user': user})

###################################### 
#              TAGS 
######################################

@login_required
def tags(request):
	context = {
		'tags': Tag.objects.all()
	}
	return render(request, 'NewEra/tags.html', context)

@login_required
def create_tag(request):
	context = {}
	form = TagForm()
	context['form'] = form
	context['action'] = 'Create'

	if request.method == 'POST':
		tag = Tag()
		form = TagForm(request.POST, instance=tag)
		
		if form.is_valid():
			form.save()
			tag.save()

			messages.success(request, 'Tag successfully created!')

			return redirect('Tags')
	else:
		tag = TagForm()

	return render(request, 'NewEra/edit_tag.html', context)

@login_required
def edit_tag(request, id):
	tag = get_object_or_404(Tag, id=id)

	if request.method == "POST":
		form = TagForm(request.POST, instance=tag)
    
		if form.is_valid():
			form.save()
			tag.save()

			messages.success(request, '{} successfully edited.'.format(tag.name))
			return redirect('Tags')
	else:
		form = TagForm(instance=tag)
	return render(request, 'NewEra/edit_tag.html', {'form': form, 'tag': tag, 'action': 'Edit'})

@login_required
def delete_tag(request, id):
	tag = get_object_or_404(Tag, id=id)

	if request.method == 'POST':
		tag.delete()
		messages.success(request, '{} successfully deleted.'.format(tag.name))
		return redirect('Tags')

	return render(request, 'NewEra/delete_tag.html', {'tag': tag})

###################################### 
# KPI SPREADSHEET EXPORT 
######################################

# Export data on resources and referrals
@login_required
def export_data(request):
	# Get resources
	resources = Resource.objects.all().filter(is_active=True)
	
	# Define the workbook and sheet
	wb = Workbook()
	ws = wb.active

	# Set the bold font
	bold = Font(bold=True)

	# Create Header row
	ws['A1'].font = bold
	ws['B1'].font = bold
	ws['C1'].font = bold
	ws['D1'].font = bold

	ws['A1'] = "Resource"
	ws['B1'] = "# Referrals"
	ws['C1'] = "# Accessed Referrals"
	ws['D1'] = "# Views"

	# Write resources
	for r in resources:
		# Get name
		name = r.name
		# Get referrals including the resource
		referrals_count = r.referrals.count()
		# Get accessed referrals
		accessed_count = r.referrals.exclude(date_accessed=None).count()
		# Get clicks
		clicks = r.clicks
		# Write to the Excel file
		ws.append([name, referrals_count, accessed_count, clicks])
	
	ws = wb.worksheets[0]
	ws.title = "Resources"
	ws.column_dimensions[get_column_letter(1)].width = 60
	ws.column_dimensions[get_column_letter(2)].width = 20
	ws.column_dimensions[get_column_letter(3)].width = 20
	ws.column_dimensions[get_column_letter(4)].width = 20


	# Export breakdown by tag

	ws1 = wb.create_sheet("By Tag")
	ws = ws1

	# Create Header row
	ws['A1'].font = bold
	ws['B1'].font = bold
	ws['C1'].font = bold
	ws['D1'].font = bold
	ws['E1'].font = bold

	ws['A1'] = "Tag"
	ws['B1'] = "# Resources"
	ws['C1'] = "# Referrals"
	ws['D1'] = "# Accessed Referrals"
	ws['E1'] = "# Views"

	# Get tags
	tags = Tag.objects.all().exclude(resource=None)

	for t in tags:
		# Get name
		name = t.name

		resources_count = 0

		referrals_count_by_tag = 0
		accessed_count = 0
		clicks = 0

		for r in resources:
			if t in r.tags.all():
				# Increment number of resources with this tag
				resources_count += 1
				# Get referrals associated with the tag
				referrals_count_by_tag += r.referrals.count()
				# Get accessed referrals by tag
				accessed_count += r.referrals.exclude(date_accessed=None).count()
				# Get clicks
				clicks += r.clicks

		# Write to the Excel file
		ws.append([name, resources_count, referrals_count_by_tag, accessed_count, clicks])
	
	ws.column_dimensions[get_column_letter(1)].width = 30
	ws.column_dimensions[get_column_letter(2)].width = 20
	ws.column_dimensions[get_column_letter(3)].width = 20
	ws.column_dimensions[get_column_letter(4)].width = 20
	ws.column_dimensions[get_column_letter(4)].width = 20


	# Export user data

	ws2 = wb.create_sheet("By User")
	ws = ws2

	# Create Header row
	ws['A1'].font = bold
	ws['B1'].font = bold
	ws['C1'].font = bold
	ws['D1'].font = bold
	ws['E1'].font = bold

	ws['A1'] = "User"
	ws['B1'] = "# Case Load Users"
	ws['C1'] = "# Referrals"
	ws['D1'] = "# Accessed Referrals"
	ws['E1'] = "Date of Last Referral"

	# Get tags
	users = User.objects.all()

	for u in users:
		case_load_count = u.get_case_load().count()
		referrals_count = u.get_referrals().count()
		accessed_referrals_count = u.get_referrals().exclude(date_accessed=None).count()
		last_referral_date = u.get_referrals().order_by('-referral_date').first()
		if last_referral_date:
			last_referral_date = last_referral_date.referral_date.strftime('%m-%d-%Y')
		else:
			last_referral_date = "No referrals made"

		# Write to the Excel file
		ws.append([str(u), case_load_count, referrals_count, accessed_referrals_count, last_referral_date])

	ws.column_dimensions[get_column_letter(1)].width = 30
	ws.column_dimensions[get_column_letter(2)].width = 20
	ws.column_dimensions[get_column_letter(3)].width = 20
	ws.column_dimensions[get_column_letter(4)].width = 20
	ws.column_dimensions[get_column_letter(4)].width = 20


	# Export referral data by phone
	ws3 = wb.create_sheet("By Referred Phone")
	ws = ws3

	# Create header row
	ws['A1'].font = bold
	ws['B1'].font = bold
	ws['C1'].font = bold
	ws['D1'].font = bold
	ws['E1'].font = bold

	ws['A1'] = "Phone"
	ws['B1'] = "Case Load User"
	ws['C1'] = "# Referrals"
	ws['D1'] = "# Accessed Referrals"
	ws['E1'] = "Date of Last Referral"

	# Get referrals
	referrals = Referral.objects.all().order_by('-referral_date')

	# Sets to keep track of phones and emails seen
	phones = set()
	emails = set()
	case_load_dict = dict()
	referrals_dict = dict()
	accessed_referrals_dict = dict()
	last_referral_dict = dict()

	for r in referrals:
		# Organize and count by phone
		if r.phone:
			# Format phone number
			if (len(r.phone) == 10):
				phone_number = "(" + r.phone[0:3] + ") " + r.phone[3:6] + "-" + r.phone[6:10]
			else:
				phone_number = r.phone[0] + " (" + r.phone[1:4] + ") " + r.phone[4:7] + "-" + r.phone[7:11]
			referrals_dict, accessed_referrals_dict, case_load_dict, last_referral_dict = export_attribute(phone_number, phones, referrals_dict, accessed_referrals_dict, case_load_dict, last_referral_dict, r)
		# Organize and count by email
		if r.email:
			referrals_dict, accessed_referrals_dict, case_load_dict, last_referral_dict = export_attribute(r.email, emails, referrals_dict, accessed_referrals_dict, case_load_dict, last_referral_dict, r)

	for p in phones:
		# Write to the Excel file
		ws.append([p, case_load_dict[p], referrals_dict[p], accessed_referrals_dict[p], last_referral_dict[p]])

	ws.column_dimensions[get_column_letter(1)].width = 30
	ws.column_dimensions[get_column_letter(2)].width = 20
	ws.column_dimensions[get_column_letter(3)].width = 20
	ws.column_dimensions[get_column_letter(4)].width = 20
	ws.column_dimensions[get_column_letter(4)].width = 20


	# Export referral data by phone
	ws4 = wb.create_sheet("By Referred Email")
	ws = ws4

	# Create header row
	ws['A1'].font = bold
	ws['B1'].font = bold
	ws['C1'].font = bold
	ws['D1'].font = bold
	ws['E1'].font = bold

	ws['A1'] = "Email"
	ws['B1'] = "Case Load User"
	ws['C1'] = "# Referrals"
	ws['D1'] = "# Accessed Referrals"
	ws['E1'] = "Date of Last Referral"

	for e in emails:
		# Write to the Excel file
		ws.append([e, case_load_dict[e], referrals_dict[e], accessed_referrals_dict[e], last_referral_dict[e]])

	ws.column_dimensions[get_column_letter(1)].width = 30
	ws.column_dimensions[get_column_letter(2)].width = 20
	ws.column_dimensions[get_column_letter(3)].width = 20
	ws.column_dimensions[get_column_letter(4)].width = 20
	ws.column_dimensions[get_column_letter(4)].width = 20

	# Save and download the Excel file
	response = HttpResponse(content_type='application/vnd.ms-excel')
	response['Content-Disposition'] = "attachment; filename=newera412_data_spreadsheet.xlsx"
	wb.save(response)

	return response

# Helper method to export attributes in a sheet
def export_attribute(attr, attr_set, referrals_dict, accessed_referrals_dict, case_load_dict, last_referral_dict, r):
	# Add email to set
	if (attr not in attr_set):
		attr_set.add(attr)
		referrals_dict[attr] = 0
		accessed_referrals_dict[attr] = 0

	# Add the case load user to the dictionary
	if (r.caseUser):
		case_load_dict[attr] = r.caseUser.get_full_name()
	else:
		case_load_dict[attr] = "-"

	# Add referrals given to this phone number
	referrals_dict[attr] += 1

	if (r.date_accessed):
		# Add accessed referrals
		accessed_referrals_dict[attr] += 1

	# Get the last referral made
	last_referral_dict[attr] = r.referral_date.strftime('%m-%d-%Y')

	return referrals_dict, accessed_referrals_dict, case_load_dict, last_referral_dict
